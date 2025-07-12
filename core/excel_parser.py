"""
Excel 檔案解析功能
"""
import os
import time
import zipfile
import xml.etree.ElementTree as ET
import re
import json
import hashlib
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.utils.exceptions import InvalidFileException
import config.settings as settings
from utils.cache import copy_to_cache
from utils.logging import get_logger

# 獲取日誌器
logger = get_logger(__name__)

def extract_external_refs(xlsx_path):
    """
    解析 Excel xlsx 中 external reference mapping: [n] -> 路徑
    """
    ref_map = {}
    try:
        with zipfile.ZipFile(xlsx_path, 'r') as z:
            rels = ET.fromstring(z.read('xl/_rels/workbook.xml.rels'))
            for rel in rels.findall('{http://schemas.openxmlformats.org/package/2006/relationships}Relationship'):
                if rel.attrib['Type'].endswith('/externalLink'):
                    target = rel.attrib['Target']
                    m = re.search(r'externalLink(\d+)\.xml', target)
                    if m:
                        num = int(m.group(1))
                        try:
                            link_xml = z.read(f'xl/{target}')
                            link_tree = ET.fromstring(link_xml)
                            book_elem = link_tree.find('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}externalBookPr')
                            if book_elem is not None:
                                path = book_elem.attrib.get('href', '')
                            else:
                                path = ''
                            ref_map[num] = path
                        except (KeyError, ET.ParseError) as e:
                            logger.debug(f"無法解析外部連結 {target}：{e}")
                            ref_map[num] = ''
                        except Exception as e:
                            logger.warning(f"處理外部連結 {target} 時發生未預期錯誤：{type(e).__name__}: {e}")
                            ref_map[num] = ''
    except FileNotFoundError:
        logger.error(f"Excel 檔案不存在：{xlsx_path}")
    except zipfile.BadZipFile:
        logger.error(f"Excel 檔案損壞或不是有效的 ZIP 格式：{xlsx_path}")
    except PermissionError:
        logger.warning(f"無權限讀取 Excel 檔案：{xlsx_path}")
    except Exception as e:
        logger.error(f"解析 Excel 外部參照時發生未預期錯誤 {xlsx_path}：{type(e).__name__}: {e}")
    return ref_map

def pretty_formula(formula, ref_map=None):
    """
    顯示 formula 時，如果有 [n]Table! 這種 external workbook reference，會顯示來源路徑
    """
    if formula is None:
        return None
    
    # 修改：處理 ArrayFormula 物件
    if isinstance(formula, ArrayFormula):
        formula_str = formula.text if hasattr(formula, 'text') else str(formula)
    else:
        formula_str = str(formula)
    
    if ref_map:
        def repl(m):
            n = int(m.group(1))
            path = ref_map.get(n, '')
            if path:
                return f"[外部檔案{n}: {path}]{m.group(0)}"
            else:
                return m.group(0)
        return re.sub(r'\[(\d+)\][A-Za-z0-9_]+!', repl, formula_str)
    else:
        return formula_str

def get_cell_formula(cell):
    """
    取得 cell 公式（不論係普通 formula or array formula），一律回傳公式字串
    """
    if cell.data_type == 'f':
        if isinstance(cell.value, ArrayFormula):
            # 修改：返回 ArrayFormula 的實際公式字符串，而不是物件
            return cell.value.text if hasattr(cell.value, 'text') else str(cell.value)
        return cell.value
    return None

def serialize_cell_value(value):
    """
    序列化儲存格值
    """
    if value is None: 
        return None
    if isinstance(value, ArrayFormula): 
        return None
    if isinstance(value, datetime): 
        return value.isoformat()
    if isinstance(value, (int, float, str, bool)): 
        return value
    return str(value)

def get_excel_last_author(path):
    """
    獲取 Excel 檔案最後修改者
    """
    try:
        wb = load_workbook(path, read_only=True)
        author = wb.properties.lastModifiedBy
        wb.close()
        del wb
        return author
    except FileNotFoundError:
        logger.error(f"Excel 檔案不存在：{path}")
        return None
    except PermissionError:
        logger.warning(f"無權限讀取 Excel 檔案：{path}")
        return None
    except InvalidFileException:
        logger.error(f"無效的 Excel 檔案格式：{path}")
        return None
    except zipfile.BadZipFile:
        logger.error(f"Excel 檔案損壞或不是有效的 ZIP 格式：{path}")
        return None
    except Exception as e:
        logger.error(f"讀取 Excel 檔案 {path} 最後修改者時發生未預期錯誤：{type(e).__name__}: {e}")
        return None

def safe_load_workbook(path, max_retry=5, delay=0.5, **kwargs):
    """
    安全載入 Excel 檔案，帶重試機制
    """
    last_err = None
    for i in range(max_retry):
        try:
            wb = load_workbook(path, **kwargs)
            return wb
        except PermissionError as e:
            logger.warning(f"載入 Excel 檔案權限被拒絕，重試 {i+1}/{max_retry}：{path}")
            last_err = e
            time.sleep(delay)
        except FileNotFoundError as e:
            logger.error(f"Excel 檔案不存在：{path}")
            last_err = e
            break
        except InvalidFileException as e:
            logger.error(f"無效的 Excel 檔案格式：{path}")
            last_err = e
            break
        except zipfile.BadZipFile as e:
            logger.error(f"Excel 檔案損壞：{path}")
            last_err = e
            break
        except Exception as e:
            logger.error(f"載入 Excel 檔案時發生未預期錯誤：{path} - {type(e).__name__}: {e}")
            last_err = e
            break
    
    logger.error(f"無法載入 Excel 檔案 {path}，已重試 {max_retry} 次")
    raise last_err

def dump_excel_cells_with_timeout(path, show_sheet_detail=True, silent=False):
    """
    提取 Excel 檔案中的所有儲存格數據
    """
    # 更新全局變數
    settings.current_processing_file = path
    settings.processing_start_time = time.time()
    
    wb = None
    try:
        if not silent: 
            print(f"   📊 檔案大小: {os.path.getsize(path)/(1024*1024):.1f} MB")
        
        local_path = copy_to_cache(path, silent=silent)
        
        # 複製完 sleep 一下，減race condition
        time.sleep(0.2)
        
        read_only_mode = True
        if not silent: 
            print(f"   🚀 讀取模式: read_only={read_only_mode}, data_only=False")
        
        wb = safe_load_workbook(local_path, read_only=read_only_mode, data_only=False)
        result = {}
        worksheet_count = len(wb.worksheets)
        
        if not silent and show_sheet_detail: 
            print(f"   📋 工作表數量: {worksheet_count}")
        
        for idx, ws in enumerate(wb.worksheets, 1):
            cell_count = 0
            ws_data = {}
            
            if ws.max_row > 1 or ws.max_column > 1:
                for row in ws.iter_rows():
                    for cell in row:
                        fstr = get_cell_formula(cell)
                        vstr = serialize_cell_value(cell.value)
                        if fstr is not None or vstr is not None:
                            ws_data[cell.coordinate] = {"formula": fstr, "value": vstr}
                            cell_count += 1
            
            if show_sheet_detail and not silent: 
                print(f"      處理工作表 {idx}/{worksheet_count}: {ws.title}（{cell_count} 有資料 cell）")
            
            if ws_data: 
                result[ws.title] = ws_data
        
        wb.close()
        wb = None
        
        if not silent and show_sheet_detail: 
            print(f"   ✅ Excel 讀取完成")
        
        return result
        
    except FileNotFoundError:
        logger.error(f"Excel 檔案不存在：{path}")
        if not silent: 
            print(f"   ❌ Excel 讀取失敗：檔案不存在")
        return None
    except PermissionError:
        logger.warning(f"無權限讀取 Excel 檔案：{path}")
        if not silent: 
            print(f"   ❌ Excel 讀取失敗：權限被拒絕")
        return None
    except InvalidFileException:
        logger.error(f"無效的 Excel 檔案格式：{path}")
        if not silent: 
            print(f"   ❌ Excel 讀取失敗：檔案格式無效")
        return None
    except zipfile.BadZipFile:
        logger.error(f"Excel 檔案損壞：{path}")
        if not silent: 
            print(f"   ❌ Excel 讀取失敗：檔案損壞")
        return None
    except Exception as e:
        logger.error(f"讀取 Excel 檔案時發生未預期錯誤 {path}：{type(e).__name__}: {e}")
        if not silent: 
            print(f"   ❌ Excel 讀取失敗：{type(e).__name__}: {e}")
        return None
    finally:
        if wb: 
            wb.close()
            del wb
        
        # 重置全局變數
        settings.current_processing_file = None
        settings.processing_start_time = None

def hash_excel_content(cells_dict):
    """
    計算 Excel 內容的雜湊值
    """
    if cells_dict is None: 
        return None
    
    try:
        content_str = json.dumps(cells_dict, sort_keys=True, ensure_ascii=False)
        return hashlib.md5(content_str.encode('utf-8')).hexdigest()
    except (TypeError, UnicodeEncodeError) as e:
        logger.error(f"計算 Excel 內容雜湊值時發生編碼錯誤：{e}")
        return None
    except Exception as e:
        logger.error(f"計算 Excel 內容雜湊值時發生未預期錯誤：{type(e).__name__}: {e}")
        return None