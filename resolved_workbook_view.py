import openpyxl
import os
import re

# 輔助函數：從工作簿中獲取外部連結映射
def _get_external_link_map(workbook):
    external_link_map = {}
    if hasattr(workbook, '_external_links') and workbook._external_links:
        for i, link in enumerate(workbook._external_links):
            if hasattr(link, 'file_link') and hasattr(link.file_link, 'target'):
                target_path = link.file_link.target
                excel_formatted_path_part = ""
                if target_path.startswith('file:///'):
                    actual_path = target_path[len('file:///'):]
                    actual_path = actual_path.replace('\\', '\\\\')
                    actual_path = actual_path.replace('/', '\\\\')

                    dirname = os.path.dirname(actual_path)
                    basename = os.path.basename(actual_path)
                    excel_formatted_path_part = f"'{dirname}\\\\[{basename}]'"
                else:
                    excel_formatted_path_part = f"[{target_path}]"
                external_link_map[str(i + 1)] = excel_formatted_path_part
    return external_link_map

# 輔助函數：解析公式字串
def _resolve_formula_string(formula_str, external_link_map):
    for index_str, formatted_path in external_link_map.items():
        formula_str = re.sub(r'\\[{}\\]'.format(re.escape(index_str)), formatted_path, formula_str)
    return formula_str


class ResolvedCellView:
    """
    包裝 openpyxl.Cell 物件，並在存取其值時解析外部連結。
    同時支援修改值和部分常用屬性。
    """
    def __init__(self, openpyxl_cell, external_link_map):
        self._cell = openpyxl_cell
        self._external_link_map = external_link_map

    @property
    def coordinate(self):
        return self._cell.coordinate

    @property
    def row(self):
        return self._cell.row

    @property
    def column(self):
        return self._cell.column

    @property
    def value(self):
        if self._cell.data_type == 'f':
            # 如果是公式，則解析並返回
            return _resolve_formula_string(self._cell.value, self._external_link_map)
        else:
            # 否則直接返回原始值
            return self._cell.value

    @value.setter
    def value(self, new_value):
        # 當設置值時，直接設置到底層的 openpyxl.Cell 物件
        self._cell.value = new_value

    @property
    def data_type(self):
        return self._cell.data_type

    # 常用樣式屬性的傳遞 (只讀取，如需修改需更複雜的包裝)
    @property
    def font(self):
        return self._cell.font

    @property
    def fill(self):
        return self._cell.fill

    @property
    def border(self):
        return self._cell.border

    @property
    def alignment(self):
        return self._cell.alignment

    @property
    def number_format(self):
        return self._cell.number_format

    @number_format.setter
    def number_format(self, value):
        self._cell.number_format = value


class ResolvedSheetView:
    """
    包裝 openpyxl.Worksheet 物件，並提供方法來獲取 ResolvedCellView 物件。
    同時支援部分常用修改操作。
    """
    def __init__(self, openpyxl_sheet, external_link_map):
        self._sheet = openpyxl_sheet
        self._external_link_map = external_link_map

    @property
    def title(self):
        return self._sheet.title

    @property
    def min_row(self):
        return self._sheet.min_row

    @property
    def max_row(self):
        return self._sheet.max_row

    @property
    def min_column(self):
        return self._sheet.min_column

    @property
    def max_column(self):
        return self._sheet.max_column

    def iter_rows(self, min_row=None, max_row=None, min_col=None, max_col=None):
        # 遍歷原始工作表的行，並返回 ResolvedCellView 物件
        for row in self._sheet.iter_rows(min_row, max_row, min_col, max_col):
            yield tuple(ResolvedCellView(cell, self._external_link_map) for cell in row)

    def __getitem__(self, key):
        # 允許通過座標訪問儲存格，例如 sheet_view['A1']
        cell = self._sheet[key]
        return ResolvedCellView(cell, self._external_link_map)

    def cell(self, row, column, value=None):
        # 模仿 openpyxl.Worksheet.cell() 方法，返回 ResolvedCellView
        original_cell = self._sheet.cell(row=row, column=column, value=value)
        return ResolvedCellView(original_cell, self._external_link_map)

    def append(self, iterable):
        # 模仿 openpyxl.Worksheet.append()，直接操作底層工作表
        self._sheet.append(iterable)

    def insert_rows(self, idx, amount=1):
        # 模仿 openpyxl.Worksheet.insert_rows()
        self._sheet.insert_rows(idx, amount)

    def delete_rows(self, idx, amount=1):
        # 模仿 openpyxl.Worksheet.delete_rows()
        self._sheet.delete_rows(idx, amount)

    def insert_cols(self, idx, amount=1):
        # 模仿 openpyxl.Worksheet.insert_cols()
        self._sheet.insert_cols(idx, amount)

    def delete_cols(self, idx, amount=1):
        # 模仿 openpyxl.Worksheet.delete_cols()
        self._sheet.delete_cols(idx, amount)

    def merge_cells(self, range_string=None, start_row=None, start_column=None, end_row=None, end_column=None):
        # 模仿 openpyxl.Worksheet.merge_cells()
        self._sheet.merge_cells(range_string, start_row, start_column, end_row, end_column)

    def unmerge_cells(self, range_string=None, start_row=None, start_column=None, end_row=None, end_column=None):
        # 模仿 openpyxl.Worksheet.unmerge_cells()
        self._sheet.unmerge_cells(range_string, start_row, start_column, end_row, end_column)

    @property
    def column_dimensions(self):
        return self._sheet.column_dimensions

    @property
    def row_dimensions(self):
        return self._sheet.row_dimensions


class ResolvedWorkbookView:
    """
    包裝 openpyxl.Workbook 物件，並提供類似的介面，但其儲存格值會解析外部連結。
    同時支援部分常用修改操作和儲存。
    """
    def __init__(self, openpyxl_workbook):
        self._workbook = openpyxl_workbook
        self._external_link_map = _get_external_link_map(openpyxl_workbook)

    @property
    def active(self):
        # 返回活動工作表的 ResolvedSheetView
        return ResolvedSheetView(self._workbook.active, self._external_link_map)

    @property
    def sheetnames(self):
        return self._workbook.sheetnames

    def __getitem__(self, key):
        # 允許通過名稱訪問工作表，例如 workbook_view['Sheet1']
        sheet = self._workbook[key]
        return ResolvedSheetView(sheet, self._external_link_map)

    def create_sheet(self, title=None, index=None):
        # 模仿 openpyxl.Workbook.create_sheet()，並返回 ResolvedSheetView
        new_sheet = self._workbook.create_sheet(title=title, index=index)
        return ResolvedSheetView(new_sheet, self._external_link_map)

    def remove(self, worksheet):
        # 模仿 openpyxl.Workbook.remove()，接受原始 openpyxl.Worksheet 或 ResolvedSheetView
        if isinstance(worksheet, ResolvedSheetView):
            self._workbook.remove(worksheet._sheet)
        else:
            self._workbook.remove(worksheet)

    def remove_sheet(self, worksheet):
        # 兼容舊版 openpyxl 的 remove_sheet 方法
        self.remove(worksheet)

    def get_sheet_by_name(self, name):
        # 模仿 openpyxl.Workbook.get_sheet_by_name()，並返回 ResolvedSheetView
        sheet = self._workbook.get_sheet_by_name(name)
        if sheet:
            return ResolvedSheetView(sheet, self._external_link_map)
        return None

    def save(self, filename):
        # 模仿 openpyxl.Workbook.save()，將底層工作簿儲存到檔案
        self._workbook.save(filename)


def my_load_workbook(file_path):
    """
    載入 Excel 檔案，並返回一個 ResolvedWorkbookView 物件。
    這個物件的儲存格值會自動解析外部連結，並提供部分修改和儲存功能。
    """
    workbook = openpyxl.load_workbook(file_path, data_only=False)
    return ResolvedWorkbookView(workbook)


# 範例使用
if __name__ == "__main__":
    # 創建一個測試檔案
    test_file_path = r"C:\Users\user\Desktop\openpyxl\test_modified.xlsx"
    original_test_file = r"C:\Users\user\Desktop\openpyxl\test.xlsx"

    # 為了測試修改功能，我們需要先複製一份 test.xlsx
    # 確保 test.xlsx 存在
    if not os.path.exists(original_test_file):
        print(f"錯誤: 找不到原始測試檔案 {original_test_file}。請確保它存在。")
    else:
        # 複製檔案以避免修改原始檔案
        import shutil
        shutil.copyfile(original_test_file, test_file_path)
        print(f"已複製 {original_test_file} 到 {test_file_path} 進行測試。\n")

        print(f"--- 使用 my_load_workbook 載入 {test_file_path} --- ")
        resolved_wb = my_load_workbook(test_file_path)
        resolved_sheet = resolved_wb.active

        print(f"偵測到的已使用範圍: {resolved_sheet.min_row}:{resolved_sheet.max_row},{resolved_sheet.min_column}:{resolved_sheet.max_column}")
        print(f"正在讀取第 {resolved_sheet.min_row} 行到第 {resolved_sheet.max_row} 行，以及第 {resolved_sheet.min_column} 列到第 {resolved_sheet.max_column} 列\n")

        # 讀取並顯示解析後的公式
        print("--- 讀取解析後的儲存格內容 ---")
        for row_view in resolved_sheet.iter_rows():
            for cell_view in row_view:
                if cell_view.value is None:
                    continue
                
                if cell_view.data_type == 'f':
                    print(f"{cell_view.coordinate}: Formula = {cell_view.value}")
                else:
                    print(f"{cell_view.coordinate}: Value = '{cell_view.value}'")

        # 測試修改功能
        print("\n--- 測試修改功能 ---")
        # 修改一個儲存格的值
        resolved_sheet['A1'].value = "新的值"
        print(f"A1 的新值: {resolved_sheet['A1'].value}")

        # 添加一行
        resolved_sheet.append(["新行1", "新行2", "新行3"])
        print(f"已添加一行。最後一行內容: {resolved_sheet.cell(row=resolved_sheet.max_row, column=1).value}, {resolved_sheet.cell(row=resolved_sheet.max_row, column=2).value}")

        # 創建一個新工作表
        new_sheet = resolved_wb.create_sheet("新的工作表")
        new_sheet.append(["新表頭", "新數據"])
        print(f"已創建新工作表: {new_sheet.title}")

        # 儲存修改後的檔案
        resolved_wb.save(test_file_path)
        print(f"\n已將修改儲存到 {test_file_path}")

        # 重新載入並驗證修改
        print("\n--- 重新載入並驗證修改 ---")
        reloaded_wb = my_load_workbook(test_file_path)
        reloaded_sheet = reloaded_wb.active
        print(f"重新載入後 A1 的值: {reloaded_sheet['A1'].value}")
        print(f"重新載入後最後一行內容: {reloaded_sheet.cell(row=reloaded_sheet.max_row, column=1).value}, {reloaded_sheet.cell(row=reloaded_sheet.max_row, column=2).value}")
        print(f"重新載入後新工作表是否存在: {"新的工作表" in reloaded_wb.sheetnames}")