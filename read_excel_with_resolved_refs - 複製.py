import openpyxl
import os
import re

def read_excel_content_resolved_refs(file_path):
    """
    讀取 Excel 檔案內容，並將公式中的外部連結索引替換為實際檔案路徑。

    Args:
        file_path (str): Excel 檔案的絕對路徑。

    Returns:
        list: 包含每個儲存格資訊的字典列表。
              每個字典包含 'coordinate', 'type' ('value' 或 'formula'), 'content'。
    """
    processed_cells = []
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=False)
        sheet = workbook.active

        min_row = sheet.min_row
        max_row = sheet.max_row
        min_col = sheet.min_column
        max_col = sheet.max_column

        # 建立外部連結索引到完整 Excel 格式路徑的映射
        external_link_map = {}
        if hasattr(workbook, '_external_links') and workbook._external_links:
            for i, link in enumerate(workbook._external_links):
                if hasattr(link, 'file_link') and hasattr(link.file_link, 'target'):
                    target_path = link.file_link.target
                    excel_formatted_path_part = ""
                    if target_path.startswith('file:///'):
                        # 絕對路徑: 'C:\Users\user\Desktop\openpyxl\[B1.xlsx]' 格式
                        # 移除 'file:///' 前綴
                        actual_path = target_path[len('file:///'):]
                        # 將所有反斜線替換為雙反斜線，以避免 Python 字串轉義問題
                        actual_path = actual_path.replace('\\', '\\\\')
                        # 將所有正斜線替換為雙反斜線 (如果存在)
                        actual_path = actual_path.replace('/', '\\\\')

                        dirname = os.path.dirname(actual_path)
                        basename = os.path.basename(actual_path)
                        
                        # 構建 Excel 公式所需的路徑字串
                        # 確保路徑中的反斜線在 Excel 公式中顯示為單個反斜線
                        # 例如: 'C:\Users\user\Desktop\openpyxl\[B1.xlsx]'
                        excel_formatted_path_part = f"'{dirname}\\\\[{basename}]'"
                    else:
                        # 相對路徑: [B.xlsx] 格式
                        excel_formatted_path_part = f"[{target_path}]"
                    external_link_map[str(i + 1)] = excel_formatted_path_part

        for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                if cell.value is None:
                    continue
                    
                cell_info = {
                    'coordinate': cell.coordinate,
                    'type': 'value',
                    'content': str(cell.value)
                }

                if cell.data_type == 'f':
                    formula_str = cell.value
                    # 使用正則表達式替換 [索引] 為完整的路徑
                    for index_str, formatted_path in external_link_map.items():
                        formula_str = re.sub(r'\[{}\]'.format(re.escape(index_str)), formatted_path, formula_str)
                    
                    cell_info['type'] = 'formula'
                    cell_info['content'] = formula_str
                
                processed_cells.append(cell_info)

    except FileNotFoundError:
        processed_cells.append({'error': f"錯誤: 找不到檔案 {file_path}"})
    except Exception as e:
        processed_cells.append({'error': f"發生意外錯誤: {e}"})
    
    return processed_cells

# 範例使用
if __name__ == "__main__":
    excel_file = r"C:\Users\user\Desktop\openpyxl\test.xlsx"
    result = read_excel_content_resolved_refs(excel_file)
    
    print(f"\n--- 處理後的 {excel_file} 內容 --- ")
    for cell_data in result:
        if 'error' in cell_data:
            print(cell_data['error'])
        else:
            print(f"{cell_data['coordinate']}: {cell_data['type']} = {cell_data['content']}")
